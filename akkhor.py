import openpyxl as xl
import re as regex

workbook = xl.Workbook()
sheet = workbook.active
workbook.save("akkhor.xlsx")

question_flag = False
ko_option_flag = False
kho_option_flag = False
go_option_flag = False
gho_option_flag = False
answer_flag = False
definition_flag = False
question = ""
option = ""
answer = ""
definition = ""
# options
ko = ""
kho = ""
go = ""
gho = ""

row = 0
sheetCount = 0

with open('2.txt', 'r') as file:
    for line in file:
        # print(line)
        # if line == "৩.১ এবং ৩.২ অর্থ সমমূল্যের ধারণা ও গুরুত্ব:\n":
        #     break
        # if regex.search("([০-৯].){2,3}", line):
        #     continue
        if regex.search("([অ-য়]+ অধ্যায়)|([০-৯৯][অ-য়] অধ্যায়)", line):
            # reset
            if answer_flag:
                answer_flag = False
                print(answer)
                if answer == " (ক)":
                    sheet.cell(row, 6, ko)
                if answer == " (খ)":
                    sheet.cell(row, 6, kho)
                if answer == " (গ)":
                    sheet.cell(row, 6, go)
                if answer == " (ঘ)":
                    sheet.cell(row, 6, gho)
                answer = ""
            if definition_flag:
                definition_flag = False
                print(definition)
                sheet.cell(row, 7, definition)
                definition = ""
            if ko_option_flag:
                ko_option_flag = False
                print(option)
                ko = option
                sheet.cell(row, 2, option)
                option = ""
            if kho_option_flag:
                kho_option_flag = False
                print(option)
                kho = option
                sheet.cell(row, 3, option)
                option = ""
            if go_option_flag:
                go_option_flag = False
                print(option)
                go = option
                sheet.cell(row, 4, option)
                option = ""
            if gho_option_flag:
                gho_option_flag = False
                print(option)
                gho = option
                sheet.cell(row, 5, option)
                option = ""
            sheetCount += 1
            sheetName = "chapter " + str(sheetCount)
            workbook.create_sheet(sheetName)
            workbook.active = workbook[sheetName]
            sheet = workbook.active
            workbook.save("akkhor.xlsx")
            row = 0
        # if row == 4:
        #     break
        for word in line.split():

            #  [q o1 o2 o3 o4 ans def]
            # question
            if regex.search("[০-৯৯]।", word):
                # print(word)
                # raise a question flag
                # reset
                if answer_flag:
                    answer_flag = False
                    print(answer)
                    if answer == " (ক)":
                        sheet.cell(row, 6, ko)
                    if answer == " (খ)":
                        sheet.cell(row, 6, kho)
                    if answer == " (গ)":
                        sheet.cell(row, 6, go)
                    if answer == " (ঘ)":
                        sheet.cell(row, 6, gho)
                    answer = ""
                if definition_flag:
                    definition_flag = False
                    print(definition)
                    sheet.cell(row, 7, definition)
                    definition = ""
                if ko_option_flag:
                    ko_option_flag = False
                    print(option)
                    ko = option
                    sheet.cell(row, 2, option)
                    option = ""
                if kho_option_flag:
                    kho_option_flag = False
                    print(option)
                    kho = option
                    sheet.cell(row, 3, option)
                    option = ""
                if go_option_flag:
                    go_option_flag = False
                    print(option)
                    go = option
                    sheet.cell(row, 4, option)
                    option = ""
                if gho_option_flag:
                    gho_option_flag = False
                    print(option)
                    gho = option
                    sheet.cell(row, 5, option)
                    option = ""
                row += 1
                question_flag = True
                continue
            #

            if not answer_flag:
                if word == "(ক)":
                    question_flag = False
                    print(question)
                    sheet.cell(row, 1, question)
                    question = ""
                    ko_option_flag = True
                    continue

                    # print("ko")

                if word == "(খ)":
                    # print("kho")
                    ko_option_flag = False
                    print(option)
                    ko = option
                    sheet.cell(row, 2, option)
                    option = ""
                    kho_option_flag = True
                    continue

                if word == "(গ)":
                    kho_option_flag = False
                    print(option)
                    kho = option
                    sheet.cell(row, 3, option)
                    option = ""
                    go_option_flag = True
                    continue

                if word == "(ঘ)":
                    go_option_flag = False
                    print(option)
                    go = option
                    sheet.cell(row, 4, option)
                    option = ""
                    gho_option_flag = True
                    continue

            # answer
            if word == "উত্তরঃ":
                if gho_option_flag:
                    gho_option_flag = False
                    print(option)
                    gho = option
                    sheet.cell(row, 5, option)
                    option = ""
                if definition_flag:
                    definition_flag = False
                    print(definition)
                    sheet.cell(row, 7, definition)
                    definition = ""
                if question_flag:
                    question_flag = False
                    print(question)
                    sheet.cell(row, 1, question)
                    question = ""
                answer_flag = True
                continue

            # definition
            if word == "ব্যাখ্যাঃ" or word == "ব্যাখাঃ" or word == "নোট":
                if gho_option_flag:
                    gho_option_flag = False
                    print(option)
                    gho = option
                    sheet.cell(row, 5, option)
                    option = ""
                if answer_flag:
                    answer_flag = False
                    print(answer)
                    if answer == " (ক)":
                        sheet.cell(row, 6, ko)
                    if answer == " (খ)":
                        sheet.cell(row, 6, kho)
                    if answer == " (গ)":
                        sheet.cell(row, 6, go)
                    if answer == " (ঘ)":
                        sheet.cell(row, 6, gho)
                    answer = ""
                definition_flag = True
                continue

            # add to doc file
            if question_flag:
                question += " " + word
            if ko_option_flag:
                option += " " + word
            if kho_option_flag:
                option += " " + word
            if go_option_flag:
                option += " " + word
            if gho_option_flag:
                option += " " + word
            if answer_flag:
                answer += " " + word
            if definition_flag:
                definition += " " + word

workbook.save("akkhor.xlsx")
